"""GeneradorArchivo: participante externo del diagrama de CU-04.

Convierte el Map de metricas en un archivo PDF o Excel descargable.
Usa las librerias que el documento lista como Servicios Externos:
openpyxl (.xlsx) y reportlab (.pdf).
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from backend.core.config import settings
from backend.core.tiempo import ahora
from backend.services.errors import FormatoInvalido

FORMATOS_SOPORTADOS = {"excel", "pdf"}

# Paleta del sistema, para que el reporte se vea como la aplicacion.
_ROSA = "E8489E"
_ROSA_CLARO = "FDEEF4"


class GeneradorArchivo:
    @staticmethod
    def exportar(metricas: dict, formato: str, periodo: str) -> Path:
        """`exportar(metricas, formato)` del diagrama.

        Devuelve la ruta del archivo generado.
        """
        formato = (formato or "").lower().strip()

        # alt [formato invalido]
        if formato not in FORMATOS_SOPORTADOS:
            raise FormatoInvalido(
                f"Formato no soportado: '{formato}'",
                extra={"formatosValidos": sorted(FORMATOS_SOPORTADOS)},
            )

        marca = ahora().strftime("%Y%m%d-%H%M%S")
        extension = "xlsx" if formato == "excel" else "pdf"
        destino = settings.REPORTES_DIR / f"reporte-productividad-{marca}.{extension}"

        if formato == "excel":
            GeneradorArchivo._generarExcel(metricas, periodo, destino)
        else:
            GeneradorArchivo._generarPDF(metricas, periodo, destino)

        return destino

    # ------------------------------------------------------------------ Excel
    @staticmethod
    def _generarExcel(metricas: dict, periodo: str, destino: Path) -> None:
        """RS-02: Excel con tiempo promedio, habitaciones por persona y eficiencia."""
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Productividad"

        titulo = Font(bold=True, size=14, color=_ROSA)
        cabecera = Font(bold=True, color="FFFFFF")
        relleno = PatternFill("solid", fgColor=_ROSA)

        hoja["A1"] = "Cattleya-Flow - Reporte de Productividad"
        hoja["A1"].font = titulo
        hoja["A2"] = f"Periodo: {periodo}"
        hoja["A3"] = f"Generado: {ahora().strftime('%d/%m/%Y %H:%M')}"

        # Resumen global
        hoja["A5"] = "RESUMEN GENERAL"
        hoja["A5"].font = Font(bold=True)
        resumen = [
            ("Habitaciones limpiadas", metricas["totalHabitaciones"]),
            ("Tiempo promedio (min)", metricas["tiempoPromedio"]),
            ("Personal activo", metricas["personalActivo"]),
            ("Eficiencia del equipo (%)", metricas["eficienciaGlobal"]),
        ]
        for i, (etiqueta, valor) in enumerate(resumen, start=6):
            hoja[f"A{i}"] = etiqueta
            hoja[f"B{i}"] = valor

        # Detalle por empleado
        fila = 6 + len(resumen) + 1
        hoja[f"A{fila}"] = "DETALLE POR EMPLEADO"
        hoja[f"A{fila}"].font = Font(bold=True)
        fila += 1

        columnas = ["Empleado", "Habitaciones", "Tiempo promedio (min)", "Eficiencia (%)"]
        for col, nombre in enumerate(columnas, start=1):
            celda = hoja.cell(row=fila, column=col, value=nombre)
            celda.font = cabecera
            celda.fill = relleno
            celda.alignment = Alignment(horizontal="center")

        for empleado in metricas["empleados"]:
            fila += 1
            hoja.cell(row=fila, column=1, value=empleado["name"])
            hoja.cell(row=fila, column=2, value=empleado["rooms"])
            hoja.cell(row=fila, column=3, value=empleado["avgMin"])
            hoja.cell(row=fila, column=4, value=empleado["eff"])

        for columna, ancho in zip("ABCD", (28, 14, 22, 14)):
            hoja.column_dimensions[columna].width = ancho

        libro.save(destino)

    # -------------------------------------------------------------------- PDF
    @staticmethod
    def _generarPDF(metricas: dict, periodo: str, destino: Path) -> None:
        documento = SimpleDocTemplate(
            str(destino),
            pagesize=A4,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )
        estilos = getSampleStyleSheet()
        elementos = []

        elementos.append(Paragraph("Cattleya-Flow", estilos["Title"]))
        elementos.append(Paragraph("Reporte de Productividad del Personal", estilos["Heading2"]))
        elementos.append(Paragraph(f"Periodo: {periodo}", estilos["Normal"]))
        elementos.append(
            Paragraph(
                f"Generado: {ahora().strftime('%d/%m/%Y %H:%M')}", estilos["Normal"]
            )
        )
        elementos.append(Spacer(1, 0.8 * cm))

        # Resumen
        elementos.append(Paragraph("Resumen general", estilos["Heading3"]))
        tablaResumen = Table(
            [
                ["Habitaciones limpiadas", str(metricas["totalHabitaciones"])],
                ["Tiempo promedio", f"{metricas['tiempoPromedio']} min"],
                ["Personal activo", str(metricas["personalActivo"])],
                ["Eficiencia del equipo", f"{metricas['eficienciaGlobal']}%"],
            ],
            colWidths=[8 * cm, 6 * cm],
        )
        tablaResumen.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{_ROSA_CLARO}")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#F3D9E6")),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("PADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        elementos.append(tablaResumen)
        elementos.append(Spacer(1, 0.8 * cm))

        # Detalle
        elementos.append(Paragraph("Detalle por empleado", estilos["Heading3"]))
        filas = [["Empleado", "Habitaciones", "Tiempo prom.", "Eficiencia"]]
        for empleado in metricas["empleados"]:
            filas.append(
                [
                    empleado["name"],
                    str(empleado["rooms"]),
                    f"{empleado['avgMin']} min",
                    f"{empleado['eff']}%",
                ]
            )

        if len(filas) == 1:
            filas.append(["Sin datos en el periodo", "-", "-", "-"])

        tablaDetalle = Table(filas, colWidths=[7 * cm, 3 * cm, 3.5 * cm, 3 * cm])
        tablaDetalle.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_ROSA}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#F3D9E6")),
                    ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("PADDING", (0, 0), (-1, -1), 5),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor(f"#{_ROSA_CLARO}")],
                    ),
                ]
            )
        )
        elementos.append(tablaDetalle)

        documento.build(elementos)
